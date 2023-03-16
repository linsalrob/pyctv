"""
Parse the VMR file and return a set (list?) of VMR objects
"""

import sys
from openpyxl import load_workbook
from .update_vmr import current_vmr
from .vmr import VMR
from .colours import message
from .helper import excel_columns
from .vmr_error import VMRBadlyFormatted

__author__ = 'Rob Edwards'

vmr_list = []


def parse_vmr(filename: str = None, verbose: bool = False) -> list:
    """
    Parse a VMR file and create a list of VMR objects
    :param filename: the filename of the VMR file to parse. If blank, we'll get it from update_vmr
    :param verbose: more output
    :return: a list of the VMR objects
    """

    global vmr_list

    if not filename:
        filename = current_vmr(verbose=verbose)

    wb = load_workbook(filename=filename, read_only=True)
    datasheet = None
    for s in wb.sheetnames:
        if s.lower().startswith("vmr"):
            datasheet = s
    if verbose:
        message(f"Parsing excel sheet: {datasheet}", "GREEN")

    cols = excel_columns()  # this is an ordered dict of the columns, but I don't know that the order matters
    entries = []  # this is the order from this file
    for row in wb[datasheet].rows:
        if not entries:
            # this is our header row. Check to see if we have the columns we expect!
            for title in row:
                if title.value not in cols:
                    errstr = f"We found an unexpected column '{title.value}' in the first row of the VMR. Please check."
                    raise VMRBadlyFormatted(errstr)
                entries.append(cols[title.value])
            continue
        d = {}
        for i, val in enumerate(row):
            d[entries[i]] = val.value
        vmr = VMR()
        vmr.set_attributes(d)
        vmr_list.append(vmr)

    return vmr_list


def genbank2vmr(vmrs: list = None, verbose: bool = False) -> dict:
    """
    Create a mapping between known GenBank identifiers and their VMR objects
    :param vmrs: a list of the VMR objects (can be None and we'll get it)
    :param verbose: more output
    :return: tuple of dicts of the genbank -> vmr objects and refseq -> vmr
    """

    global vmr_list

    if vmrs:
        vmr_list = vmrs

    gb2vmr = {}

    if not vmr_list:
        vmr_list = parse_vmr(verbose=verbose)

    for v in vmr_list:
        for acc in v.genbank_ids:
            gb2vmr[acc] = v
    return gb2vmr


def refseq2vmr(vmrs: list = None, verbose: bool = False) -> dict:
    """
    Create a mapping between known GenBank identifiers and their VMR objects
    :param vmrs: a list of the VMR objects (can be None and we'll get it)
    :param verbose: more output
    :return: tuple of dicts of the genbank -> vmr objects and refseq -> vmr
    """

    global vmr_list

    if vmrs:
        vmr_list = vmrs

    rf2vmr = {}
    if not vmr_list:
        vmr_list = parse_vmr(verbose=verbose)

    for v in vmr_list:
        for acc in v.refseq_ids:
            rf2vmr[acc] = v
    return rf2vmr
